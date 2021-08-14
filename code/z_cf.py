##
import requests
from requests.exceptions import ConnectionError
from persiantools import characters, digits
import re
from unidecode import unidecode
from lxml import etree
from io import StringIO
import pandas as pd
import openpyxl as pyxl
import pathlib
from pathlib import Path, PurePath
import os
import shutil

from Code import z_ns as ns


parser = etree.HTMLParser()

dirs = ns.Dirs()
vif = ns.VeryImportantFiles()
cte = ns.Constants()


def load_modules_pths(py_dir_n: Path = Path()):
    pydir = Path().joinpath(py_dir_n)
    pns = list(pydir.glob('*.py'))
    pns = [PurePath(x.stem).stem for x in pns]
    main_module = [x for x in pns if x.split('_')[0] == 'a'][0]
    pns = [x for x in pns if x != main_module]
    pns.sort(key=str)
    pns = [str(x).split('.py')[0] for x in pns]
    pns = [str(x).replace('/', '.') for x in pns]
    return pns


def is_machine_connected_to_internet(host="https://google.com"):
    try:
        requests.get(host)
        print("Connected to internet!")
        return True
    except ConnectionError:
        print("No internet!")
        return False


def find_jdate(inp):
    inp = str(inp)
    inp = inp[inp.find("/") - 4: inp.find("/") + 6]
    inp = inp.replace("/", "")
    date1 = str(unidecode(inp))
    if re.match(r"\d{8}", date1):
        return int(date1)
    return -1


def return_clusters_indices(iterable_obj, clustersize=100):
    intdiv = len(iterable_obj) // clustersize
    clusters_indices = [x * clustersize for x in range(0, intdiv + 1)]
    if len(clusters_indices) > 1:
        if clusters_indices[-1] != len(iterable_obj):
            clusters_indices.append(
                    clusters_indices[-1] + len(iterable_obj) % clustersize)
    else:
        clusters_indices = [0, len(iterable_obj)]
        if clusters_indices == [0, 0]:
            clusters_indices = [0]
    print(clusters_indices)
    return clusters_indices


def convert_pubdatetime_to_int(pubdate):
    pubdate = pubdate.replace("/", "")
    pubdate = pubdate.replace(":", "")
    pubdate = pubdate.replace(" ", "")
    pubdate = unidecode(pubdate)
    if re.match(r"\d{14}", pubdate):
        return int(pubdate)
    else:
        return -1


def wos(inp):
    inp1 = str(inp)
    inp1 = inp1.replace(' ', '')
    inp1 = inp1.replace('\u200c', '')
    inp1 = inp1.replace('\u202b', '')
    inp1 = inp1.replace('\n', '')
    inp1 = inp1.replace('\r\n', '')
    inp1 = inp1.replace('\t', '')
    inp1 = inp1.replace(',', '')
    inp1 = characters.ar_to_fa(inp1)
    inp1 = digits.ar_to_fa(inp1)
    inp1 = digits.fa_to_en(inp1)

    return inp1


def make_output_dict(outputskeys, default_value=None):
    outputs_dict = {}
    for el in outputskeys:
        outputs_dict[el] = default_value
    return outputs_dict


# Removes hidden rows and cols from html code to pandas can read it correctly
def fix_html(html):
    tree1 = etree.parse(StringIO(html), parser)
    for element in tree1.xpath("//*[@hidden]"):
        element.set("rowspan", "0")
        element.set("colspan", "0")
    for element in tree1.xpath('//*[contains(@style, "display:none")]'):
        element.set("rowspan", "0")
        element.set("colspan", "0")
    fixedhtml = etree.tostring(tree1, method="html", encoding="unicode")
    return fixedhtml


def find_n_month_before(current_month, howmany=1):
    if howmany == 1:
        if current_month % 100 == 1:
            previous_month = (current_month // 100 - 1) * 100 + 12
        else:
            previous_month = current_month - 1
        return previous_month
    if howmany == 0:
        return current_month
    return find_n_month_before(find_n_month_before(current_month, 1),
                               howmany - 1)


def any_of_list_isin(srchlist: list, inp):
    inp = str(inp)
    for el1 in srchlist:
        if el1 in inp:
            return True
    return False


def find_all_locs_eq_val(dfobj: pd.DataFrame, value):
    return dfobj[dfobj.eq(value)].stack().index_lbl.values.tolist()


def read_accvalue_from_str(string):
    string1 = str(string)
    string1 = string1.replace(",", "")
    if (")" in string1) and ("(" in string1):
        string1 = string1.replace(")", "")
        string1 = string1.replace("(", "")
        return -float(string1)
    else:
        return float(string1)


def update_with_last_data(indf, lastprqpn):
    if lastprqpn.exists():
        lastdf = pd.read_parquet(lastprqpn)
        indf.update(lastdf)
    return indf


def save_df_to_xl(df: pd.DataFrame,
                  pn_suff_less: pathlib.Path,
                  index: bool = False,
                  header: bool = True,
                  max_col_length: int = 40,
                  float_format=None):
    df.to_excel(pn_suff_less.resolve().with_suffix(".xlsx"),
                header=header,
                index=index,
                float_format=float_format)
    wb = pyxl.load_workbook(filename=pn_suff_less.resolve().with_suffix(
            ".xlsx"))
    ws = wb.active
    panes = index * ws['A'] + header * ws[1]
    for cell in panes:
        cell.style = 'Pandas'
    for column in ws.columns:
        try:
            length = max(len(str(cell.value)) for cell in column)
            length = length + 2 if length + 2 <= max_col_length else max_col_length
            ws.column_dimensions[column[0].column_letter].width = length
        except AttributeError:
            pass
    wb.save(pn_suff_less.resolve().with_suffix(".xlsx"))
    wb.close()


def copytree(src, dst, symlinks=False, ignore=None):
    for item in os.listdir(src):
        s = os.path.join(src, item)
        d = os.path.join(dst, item)
        if os.path.isdir(s):
            shutil.copytree(s, d, symlinks, ignore)
        else:
            shutil.copy2(s, d)


def make_balanced_subsample(df: pd.DataFrame,
                            col2balance_across: str,
                            target_column: str):
    """ Makes the balanced subsample of df across col2balance, it finds all target values that have at least one observation for each value of col2balance
        Returns the common target values
        """
    x = col2balance_across
    y = target_column
    common_y = set(df[y].unique())
    for _, gp in df.groupby([x]).__iter__():
        common_y &= set(gp[y].unique())
    bs = df[df[y].isin(common_y)]
    return bs


def load_whole_sample():
    """return latest whole sample"""
    with open(vif.lastData, 'r') as f:
        xln = f.read()

    xl_pn = dirs.out_data / xln
    df = pd.read_excel(xl_pn, engine='openpyxl')
    return df


def load_balanced_subsample():
    """ds."""
    bs_n_pn = dirs.raw / f'{vif.bs_name}.txt'
    with open(bs_n_pn, 'r') as f:
        bs_pn = f.read()
    return pd.read_excel(bs_pn, engine='openpyxl')


def return_the_1file_pn_with_suffix(dirpn: Path, suffix):
    pns = list(dirpn.glob(f'*{suffix}'))
    pns = [x for x in pns if not "~" in x.stem]
    the_pn = pns[0]
    return the_pn


def load_dollar_cpi():
    xl_pn = return_the_1file_pn_with_suffix(dirs.in_cpi_dollar_1xl, '.xlsx')
    dc = pd.read_excel(xl_pn, engine='openpyxl')
    dcc = ns.DollarCpiCols()
    fc = ns.FormalCols()
    dc = dc.iloc[:, :3]
    dc.columns = [fc.JMonth, dcc.Dollar, dcc.CPI]
    dc = dc.convert_dtypes()
    dc = dc.set_index(fc.JMonth)
    return dc


def save_fig_as_fmt(fig, pn_suffless, fmt='eps'):
    fig.savefig(f'{pn_suffless}.{fmt}', format=fmt, dpi=1200)


def add_to_tex_data(csv_name: str,
                    data_val: pd.DataFrame or float or int or str,
                    index_lbl=None,
                    col_lbl=None, ):
    """adds data to csv data"""
    csvpn = Path(dirs.texdata).joinpath(csv_name).with_suffix('.csv')

    if csvpn.exists():
        df = pd.read_csv(csvpn)
    else:
        df = pd.DataFrame()

    if type(data_val) == pd.DataFrame:
        data_val = data_val.append(df)
        data_val = data_val[~data_val.index.duplicated()]
        df = data_val
    else:
        if csv_name == vif.vars and col_lbl is None:
            df.at[index_lbl] = data_val
        else:
            df.at[index_lbl, col_lbl] = data_val

    df = df.convert_dtypes()
    df.to_csv(csvpn)
    print(csvpn)
    return df


def main():
    pass


##
if __name__ == '__main__':
    pass
else:
    pass
    ##

##
