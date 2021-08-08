##
import requests
from requests.exceptions import ConnectionError
from persiantools import characters, digits
import re
from unidecode import unidecode
from lxml import etree
from io import StringIO
import pandas as pd
import openpyxl as pyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pathlib


try:
    from py import z_ns as ns
except ModuleNotFoundError:
    import z_ns as ns

parser = etree.HTMLParser()

def is_machine_connected_to_internet(host="https://google.com"):
    try:
        requests.get(host)
        print("Connected to internet!")
        return True
    except ConnectionError:
        print("No internet!")
        return False

def find_jdate(inp):
    inp = str(inp)
    inp = inp[inp.find("/") - 4: inp.find("/") + 6]
    inp = inp.replace("/", "")
    date1 = str(unidecode(inp))
    if re.match(r"\d{8}", date1):
        return int(date1)
    return -1

def return_clusters_indices(iterable_obj, clustersize=100):
    intdiv = len(iterable_obj) // clustersize
    clusters_indices = [x * clustersize for x in range(0, intdiv + 1)]
    if len(clusters_indices) > 1:
        if clusters_indices[-1] != len(iterable_obj):
            clusters_indices.append(
                    clusters_indices[-1] + len(iterable_obj) % clustersize)
    else:
        clusters_indices = [0, len(iterable_obj)]
        if clusters_indices == [0, 0]:
            clusters_indices = [0]
    print(clusters_indices)
    return clusters_indices

def convert_pubdatetime_to_int(pubdate):
    pubdate = pubdate.replace("/", "")
    pubdate = pubdate.replace(":", "")
    pubdate = pubdate.replace(" ", "")
    pubdate = unidecode(pubdate)
    if re.match(r"\d{14}", pubdate):
        return int(pubdate)
    else:
        return -1

def wos(inp):
    inp1 = str(inp)
    inp1 = inp1.replace(' ', '')
    inp1 = inp1.replace('\u200c', '')
    inp1 = inp1.replace('\u202b', '')
    inp1 = inp1.replace('\n', '')
    inp1 = inp1.replace('\r\n', '')
    inp1 = inp1.replace('\t', '')
    inp1 = inp1.replace(',', '')
    inp1 = characters.ar_to_fa(inp1)
    inp1 = digits.ar_to_fa(inp1)
    inp1 = digits.fa_to_en(inp1)

    return inp1

def make_output_dict(outputskeys, default_value=None):
    outputs_dict = {}
    for el in outputskeys:
        outputs_dict[el] = default_value
    return outputs_dict

# Removes hidden rows and cols from html code to pandas can read it correctly
def fix_html(html):
    tree1 = etree.parse(StringIO(html), parser)
    for element in tree1.xpath("//*[@hidden]"):
        element.set("rowspan", "0")
        element.set("colspan", "0")
    for element in tree1.xpath('//*[contains(@style, "display:none")]'):
        element.set("rowspan", "0")
        element.set("colspan", "0")
    fixedhtml = etree.tostring(tree1, method = "html", encoding = "unicode")
    return fixedhtml

def find_n_month_before(current_month, howmany=1):
    if howmany == 1:
        if current_month % 100 == 1:
            previous_month = (current_month // 100 - 1) * 100 + 12
        else:
            previous_month = current_month - 1
        return previous_month
    if howmany == 0:
        return current_month
    return find_n_month_before(find_n_month_before(current_month, 1),
                               howmany - 1)

def any_of_list_isin(srchlist: list, inp):
    inp = str(inp)
    for el1 in srchlist:
        if el1 in inp:
            return True
    return False

def find_all_locs_eq_val(dfobj: pd.DataFrame, value):
    return dfobj[dfobj.eq(value)].stack().index.values.tolist()

def read_accvalue_from_str(string):
    string1 = str(string)
    string1 = string1.replace(",", "")
    if (")" in string1) and ("(" in string1):
        string1 = string1.replace(")", "")
        string1 = string1.replace("(", "")
        return -float(string1)
    else:
        return float(string1)

def update_with_last_data(indf, lastprqpn):
    if lastprqpn.exists():
        lastdf = pd.read_parquet(lastprqpn)
        indf.update(lastdf)
    return indf

def save_df_to_xl(df: pd.DataFrame,
                  pn_suff_less: pathlib.Path,
                  index: bool = False,
                  header: bool = True,
                  max_col_length: int = 40,
                  float_format=None):
    df.to_excel(pn_suff_less.resolve().with_suffix(".xlsx"),
                header = header,
                index = index,
                float_format = float_format)
    wb = pyxl.load_workbook(filename = pn_suff_less.resolve().with_suffix(
            ".xlsx"))
    ws = wb.active
    panes = index * ws['A'] + header * ws[1]
    for cell in panes:
        cell.style = 'Pandas'
    for column in ws.columns:
        length = max(len(str(cell.value)) for cell in column)
        length = length + 2 if length + 2 <= max_col_length else max_col_length
        ws.column_dimensions[column[0].column_letter].width = length
    wb.save(pn_suff_less.resolve().with_suffix(".xlsx"))
    wb.close()
